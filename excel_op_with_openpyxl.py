# coding = utf-8
import openpyxl
from openpyxl import Workbook
from kuaidi1000_function import KuaiDi100
import json
import time

wb = Workbook()
ws = wb.active

# 定义表头
title = ['客户名称','快递单号','快递状态','最后更新时间','快递开始时间','详情','时效/天']
ws.append(title)

start_time = time.time()

#读取excel表格
customer_file_path = 'F:/EXCEL/info/Customer.xlsx'
customer_workbook = openpyxl.load_workbook(customer_file_path)

# 获取工作表的Sheet
customer_sheet = customer_workbook['海运SEA']
sheet = customer_sheet.title

#顺丰快递需要手机号验证
air = 2699
sea = 8166
phone = ''

#判断当前表格是否海运还是空运并赋值给phone
if sheet == '海运SEA':
    phone = sea
else:
    phone = air


# 获取Customer.xlsx中A2:A63列的所有客户名字和B2:B63列的相对应的快递单号
customer_names =[(customer_sheet.cell(row=i, column=1).value) for i in range(2,100)]
express_numbers = [(customer_sheet.cell(row=i, column=2).value) for i in range(2,100)]



# 将快递单号使用for循环查询快递状态 ，并append 到excel中
for express_number, customer_name in zip(express_numbers,customer_names):
        
         response = KuaiDi100().track('',express_number,phone,'','')
         #定义获取到的数据
         api_data = json.loads(response)
         #判断数据
         if 'data' in api_data and api_data['data']:
             last_data = api_data['data'][-1] #取最末尾的数据
             first_data = api_data['data'][0] #取最前面的数据
             #imfromation_last = []   最末尾的数据[]
             infomation_first = [customer_name,express_number,
                                 first_data['status'],
                                 first_data['time'],
                                 last_data['time'],
                                 first_data['context']
                             ]
             ws.append(infomation_first)
         else:
            break
       
#在G栏 填入Excel 函数来计算时效 天数为单位
for row in range(2,ws.max_row + 1):
       formula = f'=DATEDIF(E{row},D{row},"D")'
       ws[f'G{row}'] = formula

# #格式化日期和时间
formatted_time = time.strftime("%H_%M_%S", time.localtime(time.time()))
formatted_date = time.strftime("%Y_%m_%d", time.localtime(time.time()))

#定义 文件路径和文件名
output_path = 'F:/EXCEL/output/快递更新状态'+formatted_date+"_"+formatted_time+'.xlsx'

#保存文件
wb.save(output_path)

end_time = time.time()
total_time = end_time - start_time

print(f"执行程序所花时间为: {total_time}")

print('=====Done====')
